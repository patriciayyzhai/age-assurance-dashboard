#!/usr/bin/env python3
"""
Convert the H1 2026 Trust & Safety spreadsheet into the dashboard data layer.

Produces:
  data/seed/jurisdictions.json  -> all 70 markets (ISO alpha-2, region, flag)
  data/markets.json             -> market-level T&S posture records (source of truth)

Run:
  cd scripts && python convert_spreadsheet.py
"""
import json
import os
import re
from datetime import datetime, timezone

SPREADSHEET = "/Users/patzhai/Desktop/tns market dashboard/baseline_2025_updated.xlsx"
SHEET = "Reg & China Sentiment H1 2026"
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DATA = os.path.join(ROOT, "data")

# country -> (iso_alpha2, region, flag_emoji)
META = {
    "United Kingdom": ("GB", "Europe", "🇬🇧"),
    "Australia": ("AU", "Asia Pacific", "🇦🇺"),
    "Germany": ("DE", "Europe", "🇩🇪"),
    "France": ("FR", "Europe", "🇫🇷"),
    "Netherlands": ("NL", "Europe", "🇳🇱"),
    "Ireland": ("IE", "Europe", "🇮🇪"),
    "South Korea": ("KR", "Asia Pacific", "🇰🇷"),
    "Lithuania": ("LT", "Europe", "🇱🇹"),
    "Denmark": ("DK", "Europe", "🇩🇰"),
    "Sweden": ("SE", "Europe", "🇸🇪"),
    "Finland": ("FI", "Europe", "🇫🇮"),
    "Italy": ("IT", "Europe", "🇮🇹"),
    "Spain": ("ES", "Europe", "🇪🇸"),
    "Austria": ("AT", "Europe", "🇦🇹"),
    "Poland": ("PL", "Europe", "🇵🇱"),
    "Norway": ("NO", "Europe", "🇳🇴"),
    "India": ("IN", "Asia Pacific", "🇮🇳"),
    "Greece": ("GR", "Europe", "🇬🇷"),
    "Portugal": ("PT", "Europe", "🇵🇹"),
    "Singapore": ("SG", "Asia Pacific", "🇸🇬"),
    "Brazil": ("BR", "South America", "🇧🇷"),
    "Czech Republic": ("CZ", "Europe", "🇨🇿"),
    "Hungary": ("HU", "Europe", "🇭🇺"),
    "Indonesia": ("ID", "Asia Pacific", "🇮🇩"),
    "Romania": ("RO", "Europe", "🇷🇴"),
    "Vietnam": ("VN", "Asia Pacific", "🇻🇳"),
    "Malaysia": ("MY", "Asia Pacific", "🇲🇾"),
    "United States": ("US", "North America", "🇺🇸"),
    "New Zealand": ("NZ", "Asia Pacific", "🇳🇿"),
    "Slovakia": ("SK", "Europe", "🇸🇰"),
    "Bulgaria": ("BG", "Europe", "🇧🇬"),
    "Croatia": ("HR", "Europe", "🇭🇷"),
    "Japan": ("JP", "Asia Pacific", "🇯🇵"),
    "Israel": ("IL", "Middle East", "🇮🇱"),
    "Switzerland": ("CH", "Europe", "🇨🇭"),
    "Canada": ("CA", "North America", "🇨🇦"),
    "Philippines": ("PH", "Asia Pacific", "🇵🇭"),
    "Turkey": ("TR", "Middle East", "🇹🇷"),
    "South Africa": ("ZA", "Africa", "🇿🇦"),
    "Saudi Arabia": ("SA", "Middle East", "🇸🇦"),
    "United Arab Emirates": ("AE", "Middle East", "🇦🇪"),
    "Mexico": ("MX", "North America", "🇲🇽"),
    "Chile": ("CL", "South America", "🇨🇱"),
    "Nigeria": ("NG", "Africa", "🇳🇬"),
    "Egypt": ("EG", "Middle East", "🇪🇬"),
    "Qatar": ("QA", "Middle East", "🇶🇦"),
    "Ukraine": ("UA", "Europe", "🇺🇦"),
    "Costa Rica": ("CR", "North America", "🇨🇷"),
    "Colombia": ("CO", "South America", "🇨🇴"),
    "Peru": ("PE", "South America", "🇵🇪"),
    "Argentina": ("AR", "South America", "🇦🇷"),
    "Ecuador": ("EC", "South America", "🇪🇨"),
    "Dominican Republic": ("DO", "North America", "🇩🇴"),
    "Kenya": ("KE", "Africa", "🇰🇪"),
    "Ghana": ("GH", "Africa", "🇬🇭"),
    "El Salvador": ("SV", "North America", "🇸🇻"),
    "Pakistan": ("PK", "Asia Pacific", "🇵🇰"),
    "Kuwait": ("KW", "Middle East", "🇰🇼"),
    "Sri Lanka": ("LK", "Asia Pacific", "🇱🇰"),
    "Algeria": ("DZ", "Africa", "🇩🇿"),
    "Tunisia": ("TN", "Africa", "🇹🇳"),
    "Lebanon": ("LB", "Middle East", "🇱🇧"),
    "Belarus": ("BY", "Europe", "🇧🇾"),
    "Azerbaijan": ("AZ", "Asia Pacific", "🇦🇿"),
    "Venezuela": ("VE", "South America", "🇻🇪"),
    "Barbados": ("BB", "North America", "🇧🇧"),
    "Angola": ("AO", "Africa", "🇦🇴"),
    "Kazakhstan": ("KZ", "Asia Pacific", "🇰🇿"),
    "Uzbekistan": ("UZ", "Asia Pacific", "🇺🇿"),
    "Madagascar": ("MG", "Africa", "🇲🇬"),
    # --- H1 2026 expansion: near-global coverage (China deliberately excluded) ---
    "North Korea": ("KP", "Asia Pacific", "🇰🇵"),
    "Iran": ("IR", "Middle East", "🇮🇷"),
    "Russia": ("RU", "Europe", "🇷🇺"),
    "Myanmar": ("MM", "Asia Pacific", "🇲🇲"),
    "Cuba": ("CU", "North America", "🇨🇺"),
    "Laos": ("LA", "Asia Pacific", "🇱🇦"),
    "Cambodia": ("KH", "Asia Pacific", "🇰🇭"),
    "Georgia": ("GE", "Europe", "🇬🇪"),
    "Jordan": ("JO", "Middle East", "🇯🇴"),
    "Bangladesh": ("BD", "Asia Pacific", "🇧🇩"),
    "Thailand": ("TH", "Asia Pacific", "🇹🇭"),
    "Nepal": ("NP", "Asia Pacific", "🇳🇵"),
    "Oman": ("OM", "Middle East", "🇴🇲"),
    "Bahrain": ("BH", "Middle East", "🇧🇭"),
    "Ethiopia": ("ET", "Africa", "🇪🇹"),
    "Uganda": ("UG", "Africa", "🇺🇬"),
    "Zimbabwe": ("ZW", "Africa", "🇿🇼"),
    "Tanzania": ("TZ", "Africa", "🇹🇿"),
    "Rwanda": ("RW", "Africa", "🇷🇼"),
    "Cameroon": ("CM", "Africa", "🇨🇲"),
    "Morocco": ("MA", "Africa", "🇲🇦"),
    "Senegal": ("SN", "Africa", "🇸🇳"),
    "Botswana": ("BW", "Africa", "🇧🇼"),
    "Ivory Coast": ("CI", "Africa", "🇨🇮"),
    "DR Congo": ("CD", "Africa", "🇨🇩"),
    "Namibia": ("NA", "Africa", "🇳🇦"),
    "Mali": ("ML", "Africa", "🇲🇱"),
    "Burkina Faso": ("BF", "Africa", "🇧🇫"),
    "Niger": ("NE", "Africa", "🇳🇪"),
    "Chad": ("TD", "Africa", "🇹🇩"),
    "Gabon": ("GA", "Africa", "🇬🇦"),
    "Congo": ("CG", "Africa", "🇨🇬"),
    "Mozambique": ("MZ", "Africa", "🇲🇿"),
    "Zambia": ("ZM", "Africa", "🇿🇲"),
    "Malawi": ("MW", "Africa", "🇲🇼"),
    "Lesotho": ("LS", "Africa", "🇱🇸"),
    "Eswatini": ("SZ", "Africa", "🇸🇿"),
    "Mauritius": ("MU", "Africa", "🇲🇺"),
    "Guinea": ("GN", "Africa", "🇬🇳"),
    "Sierra Leone": ("SL", "Africa", "🇸🇱"),
    "Liberia": ("LR", "Africa", "🇱🇷"),
    "Togo": ("TG", "Africa", "🇹🇬"),
    "Benin": ("BJ", "Africa", "🇧🇯"),
    "Mauritania": ("MR", "Africa", "🇲🇷"),
    "Eritrea": ("ER", "Africa", "🇪🇷"),
    "Djibouti": ("DJ", "Africa", "🇩🇯"),
    "Central African Republic": ("CF", "Africa", "🇨🇫"),
    "Equatorial Guinea": ("GQ", "Africa", "🇬🇶"),
    "Somalia": ("SO", "Africa", "🇸🇴"),
    "Sudan": ("SD", "Africa", "🇸🇩"),
    "South Sudan": ("SS", "Africa", "🇸🇸"),
    "Libya": ("LY", "Africa", "🇱🇾"),
    "Kyrgyzstan": ("KG", "Asia Pacific", "🇰🇬"),
    "Tajikistan": ("TJ", "Asia Pacific", "🇹🇯"),
    "Turkmenistan": ("TM", "Asia Pacific", "🇹🇲"),
    "Armenia": ("AM", "Europe", "🇦🇲"),
    "Mongolia": ("MN", "Asia Pacific", "🇲🇳"),
    "Afghanistan": ("AF", "Asia Pacific", "🇦🇫"),
    "Bhutan": ("BT", "Asia Pacific", "🇧🇹"),
    "Brunei": ("BN", "Asia Pacific", "🇧🇳"),
    "Papua New Guinea": ("PG", "Asia Pacific", "🇵🇬"),
    "Fiji": ("FJ", "Asia Pacific", "🇫🇯"),
    "Timor-Leste": ("TL", "Asia Pacific", "🇹🇱"),
    "Iraq": ("IQ", "Middle East", "🇮🇶"),
    "Yemen": ("YE", "Middle East", "🇾🇪"),
    "Syria": ("SY", "Middle East", "🇸🇾"),
    "Iceland": ("IS", "Europe", "🇮🇸"),
    "Serbia": ("RS", "Europe", "🇷🇸"),
    "Albania": ("AL", "Europe", "🇦🇱"),
    "Bosnia and Herzegovina": ("BA", "Europe", "🇧🇦"),
    "Macedonia": ("MK", "Europe", "🇲🇰"),
    "Montenegro": ("ME", "Europe", "🇲🇪"),
    "Moldova": ("MD", "Europe", "🇲🇩"),
    "Uruguay": ("UY", "South America", "🇺🇾"),
    "Paraguay": ("PY", "South America", "🇵🇾"),
    "Bolivia": ("BO", "South America", "🇧🇴"),
    "Guyana": ("GY", "South America", "🇬🇾"),
    "Suriname": ("SR", "South America", "🇸🇷"),
    "Guatemala": ("GT", "North America", "🇬🇹"),
    "Panama": ("PA", "North America", "🇵🇦"),
    "Honduras": ("HN", "North America", "🇭🇳"),
    "Nicaragua": ("NI", "North America", "🇳🇮"),
    "Jamaica": ("JM", "North America", "🇯🇲"),
    "Trinidad and Tobago": ("TT", "North America", "🇹🇹"),
    "Belize": ("BZ", "North America", "🇧🇿"),
    "Andorra": ("AD", "Europe", "🇦🇩"),
    "Liechtenstein": ("LI", "Europe", "🇱🇮"),
    "Greenland": ("GL", "North America", "🇬🇱"),
}

# Regulatory status label -> canonical status id + display
STATUS_MAP = [
    ("passed (pre-enforcement)", "passed"),
    ("effective (state-led", "effective"),
    ("proposed / under review", "under_review"),
    ("proposed", "proposed"),
    ("enforced", "enforced"),
    ("effective", "effective"),
    ("passed", "passed"),
]

# obligation label (spreadsheet) -> obligation id (dashboard taxonomy)
OBLIGATION_MAP = {
    "Age Assurance / Gating": "age_assurance",
    "Parental Consent": "parental_consent",
    "Access Restriction": "access_restriction",
    "Design Restrictions": "design_restrictions",
    "Data Protection": "data_protection",
    "Content Moderation": "content_moderation",
    "Risk Assessment": "risk_assessment",
    "Transparency Reporting": "transparency_reporting",
    "User Empowerment / Redress": "user_empowerment",
    "AI Transparency / Content Labelling": "ai_transparency",
}

# scope token (spreadsheet) -> service_type id (dashboard taxonomy)
SCOPE_MAP = {
    "social media": "social_media",
    "gaming": "gaming",
    "ai service": "ai_service",
    "app store": "app_store",
    "streaming": "streaming",
    "adult content": "adult_content",
    "pornography": "adult_content",  # collapsed into Adult Content (dedup handles overlap)
    "online marketplace": "online_marketplace",
    "direct messaging": "direct_messaging",
    "os provider": "os_provider",
}


def band_from_score(s):
    if s is None:
        return None
    if s >= 81:
        return "severe"
    if s >= 61:
        return "high"
    if s >= 41:
        return "moderate"
    if s >= 21:
        return "low"
    return "minimal"


def map_status(label):
    l = (label or "").lower()
    for prefix, sid in STATUS_MAP:
        if l.startswith(prefix) or prefix in l:
            return sid
    return "proposed"


def map_scope(scope):
    out = []
    for tok in re.split(r"[;,]", scope or ""):
        t = tok.strip().lower()
        if not t:
            continue
        if t in SCOPE_MAP:
            out.append(SCOPE_MAP[t])
        else:
            # fuzzy contains
            for k, v in SCOPE_MAP.items():
                if k in t:
                    out.append(v)
                    break
    # dedup preserve order
    seen = set()
    return [x for x in out if not (x in seen or seen.add(x))]


def map_obligations(text):
    out = []
    for tok in (text or "").split(";"):
        t = tok.strip()
        if t in OBLIGATION_MAP:
            out.append(OBLIGATION_MAP[t])
    seen = set()
    return [x for x in out if not (x in seen or seen.add(x))]


def main():
    import openpyxl
    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    ws = wb[SHEET]

    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    markets = []
    jurisdictions = []

    r = 2
    while True:
        country = ws.cell(row=r, column=1).value
        if not country:
            break  # first empty country cell = end of data (methodology block follows)
        # stop if we've wandered into the methodology block (title row)
        if isinstance(country, str) and country.startswith("Scoring Methodology"):
            break
        if country not in META:
            raise SystemExit(f"Missing META for country: {country!r}")
        iso, region, flag = META[country]
        status_label = ws.cell(row=r, column=2).value
        scope = ws.cell(row=r, column=3).value
        # Heatmap/UI severity now reads the Gaussian display column (L / col 12).
        # The absolute researched score in col D (4) is kept in the spreadsheet
        # for reference but is intentionally NOT surfaced in the JSON/UI.
        sev_gaussian = ws.cell(row=r, column=12).value
        sev_absolute = ws.cell(row=r, column=4).value
        sev = sev_gaussian if sev_gaussian is not None else sev_absolute
        cn = ws.cell(row=r, column=6).value
        cn_trigger = ws.cell(row=r, column=8).value
        note = ws.cell(row=r, column=9).value
        url = ws.cell(row=r, column=10).value
        obl_text = ws.cell(row=r, column=11).value

        jid = iso
        jurisdictions.append({
            "id": jid, "name": country, "parent_id": None,
            "iso_alpha2": iso, "region": region, "flag_emoji": flag,
        })

        markets.append({
            "id": jid,
            "jurisdiction_id": jid,
            "country": country,
            "regulatory_status": map_status(status_label),
            "regulatory_status_label": status_label,
            "service_type_ids": map_scope(scope),
            "regulatory_severity": sev,
            "regulatory_severity_band": band_from_score(sev),
            "china_sentiment": cn,
            "china_sentiment_band": band_from_score(cn),
            "china_sentiment_trigger": cn_trigger,
            "obligation_ids": map_obligations(obl_text),
            "assessment_note": note,
            "source_url": url,
            "updated_at": now,
        })

        r += 1

    # sort jurisdictions by region then name for the filter UI
    jurisdictions.sort(key=lambda j: (j["region"], j["name"]))

    with open(os.path.join(DATA, "seed", "jurisdictions.json"), "w", encoding="utf-8") as f:
        json.dump({"version": "2.0.0", "last_updated": now, "jurisdictions": jurisdictions},
                  f, ensure_ascii=False, indent=2)

    with open(os.path.join(DATA, "markets.json"), "w", encoding="utf-8") as f:
        json.dump({"version": "2.0.0", "last_updated": now, "markets": markets},
                  f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(jurisdictions)} jurisdictions and {len(markets)} markets.")
    # quick distribution
    from collections import Counter
    sb = Counter(m["regulatory_severity_band"] for m in markets)
    cb = Counter(m["china_sentiment_band"] for m in markets)
    ob = Counter(o for m in markets for o in m["obligation_ids"])
    print("Severity bands:", dict(sb))
    print("China bands:", dict(cb))
    print("Obligation freq:", dict(ob))
    # verify all scopes mapped
    for m in markets:
        if not m["service_type_ids"]:
            print("  WARN no service types:", m["country"])
        if not m["obligation_ids"]:
            print("  WARN no obligations:", m["country"])


if __name__ == "__main__":
    main()
