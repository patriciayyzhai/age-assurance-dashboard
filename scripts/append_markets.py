#!/usr/bin/env python3
"""
Append the 88 new markets (new_markets.NEW_MARKETS) to the H1 2026 tab.

Robust approach (avoids openpyxl insert_rows + merged-cell fragility):
  1. Capture the methodology block (rows 72..max_row): values, styles, merges,
     row heights, as relative offsets from block start (row 72).
  2. Clear that region and remove its merged ranges.
  3. Write the 88 new data rows starting at row 72 (matching data formatting).
  4. Re-lay the methodology block after the new data (2-row gap), restoring
     values, styles, merges and row heights at the shifted position.
"""
import sys, os, copy
sys.path.insert(0, os.path.dirname(__file__))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from new_markets import NEW_MARKETS

SPREADSHEET = "/Users/patzhai/Desktop/tns market dashboard/baseline_2025_updated.xlsx"
SHEET = "Reg & China Sentiment H1 2026"

BAND_FILL = {
    "Severe":   ("C00000", "FFFFFF"),
    "High":     ("FF7C80", "000000"),
    "Moderate": ("FFC000", "000000"),
    "Low":      ("FFE699", "000000"),
    "Minimal":  ("C6E0B4", "000000"),
}

def band(s):
    if s >= 81: return "Severe"
    if s >= 61: return "High"
    if s >= 41: return "Moderate"
    if s >= 21: return "Low"
    return "Minimal"

def main():
    wb = openpyxl.load_workbook(SPREADSHEET)
    ws = wb[SHEET]

    # 1. detect last data row (col A non-empty, contiguous from row 2)
    last_data = 1
    for r in range(2, 400):
        if ws.cell(row=r, column=1).value is None:
            last_data = r - 1
            break
    block_start = last_data + 1            # first row of methodology region (incl. gap)
    block_end = ws.max_row
    print("last data row:", last_data, "methodology region:", block_start, "-", block_end)

    # 2. capture methodology block: per-cell value + style, relative to block_start
    captured = []  # list of (dr, col, value, style_dict)
    for r in range(block_start, block_end + 1):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None or cell.has_style:
                captured.append((
                    r - block_start, c, cell.value,
                    copy.copy(cell.font), copy.copy(cell.fill),
                    copy.copy(cell.alignment), copy.copy(cell.border),
                    cell.number_format,
                ))
    # capture merges in region (as relative row offsets)
    captured_merges = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= block_start:
            captured_merges.append((
                mr.min_row - block_start, mr.min_col,
                mr.max_row - block_start, mr.max_col,
            ))
    # capture row heights
    captured_heights = {}
    for r in range(block_start, block_end + 1):
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            captured_heights[r - block_start] = rd.height

    # 3. clear region: remove merges, then clear cells
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= block_start:
            ws.unmerge_cells(str(mr))
    for r in range(block_start, block_end + 1):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.alignment = Alignment()

    # 4. write new data rows
    n = len(NEW_MARKETS)
    insert_at = last_data + 1
    body_font = Font(name="Aptos Narrow", size=12)
    k_font = Font(name="Aptos Narrow", size=11)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i, row in enumerate(NEW_MARKETS):
        r = insert_at + i
        (country, iso, region, flag, status, scope, sev, cn,
         trig, note, url, obligations) = row
        sev_band = band(sev); cn_band = band(cn)
        obl_text = "; ".join(obligations)

        cA = ws.cell(row=r, column=1, value=country); cA.font = body_font; cA.alignment = wrap_top
        cB = ws.cell(row=r, column=2, value=status); cB.font = body_font; cB.alignment = wrap_top
        cC = ws.cell(row=r, column=3, value=scope); cC.font = body_font; cC.alignment = wrap_top
        cD = ws.cell(row=r, column=4, value=sev); cD.font = body_font; cD.alignment = center

        eb = ws.cell(row=r, column=5, value=sev_band)
        fg, fc = BAND_FILL[sev_band]
        eb.fill = PatternFill("solid", fgColor=fg)
        eb.font = Font(name="Aptos Narrow", size=12, color=fc); eb.alignment = center

        cF = ws.cell(row=r, column=6, value=cn); cF.font = body_font; cF.alignment = center

        gb = ws.cell(row=r, column=7, value=cn_band)
        fg, fc = BAND_FILL[cn_band]
        gb.fill = PatternFill("solid", fgColor=fg)
        gb.font = Font(name="Aptos Narrow", size=12, color=fc); gb.alignment = center

        cH = ws.cell(row=r, column=8, value=trig); cH.font = body_font; cH.alignment = wrap_top
        cI = ws.cell(row=r, column=9, value=note); cI.font = body_font; cI.alignment = wrap_top
        cJ = ws.cell(row=r, column=10, value=url); cJ.font = body_font; cJ.alignment = wrap_top
        cK = ws.cell(row=r, column=11, value=obl_text); cK.font = k_font; cK.alignment = left_top
        ws.row_dimensions[r].height = 51

    # 5. re-lay methodology block after data (2-row gap for visual separation)
    new_block_start = insert_at + n + 2
    for (dr, c, value, font, fill, align, border, numfmt) in captured:
        cell = ws.cell(row=new_block_start + dr, column=c)
        cell.value = value
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border
        cell.number_format = numfmt
    for (dr0, c0, dr1, c1) in captured_merges:
        ws.merge_cells(start_row=new_block_start + dr0, start_column=c0,
                       end_row=new_block_start + dr1, end_column=c1)
    for dr, h in captured_heights.items():
        ws.row_dimensions[new_block_start + dr].height = h

    wb.save(SPREADSHEET)
    print(f"Wrote {n} new markets at rows {insert_at}-{insert_at+n-1}.")
    print(f"Methodology block relocated to row {new_block_start}+.")

if __name__ == "__main__":
    main()
